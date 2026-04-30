from pathlib import Path


def parse(path: str | Path, mime_type: str = "") -> str:
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return _pdf(path)
    elif suffix in (".docx",):
        return _docx(path)
    elif suffix in (".xlsx", ".xls"):
        return _xlsx(path)
    elif suffix in (".md", ".markdown"):
        return _md(path)
    else:
        return path.read_text(encoding="utf-8", errors="ignore")


def _pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _docx(path: Path) -> str:
    from docx import Document
    doc = Document(path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            line = "\t".join(str(c) if c is not None else "" for c in row)
            if line.strip():
                lines.append(line)
    return "\n".join(lines)


def _md(path: Path) -> str:
    from markdown_it import MarkdownIt
    md = MarkdownIt()
    tokens = md.parse(path.read_text(encoding="utf-8"))
    return "\n\n".join(
        t.content for t in tokens if t.type == "inline" and t.content.strip()
    )